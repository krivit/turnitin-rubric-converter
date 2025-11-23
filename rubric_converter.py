#!/usr/bin/env python3
# -----------------------------------------------------------------------------
# Copyright (c) 2025, Pavel N. Krivitsky <p.krivitsky@unsw.edu.au>
#                     for UNSW Sydney
#
# Original script written and refined by Pavel N. Krivitsky, with iterative
# improvements and feature additions guided by GitHub Copilot
# <copilot@github.com> using GPT-4.1 backend.
# -----------------------------------------------------------------------------
#
# USAGE INSTRUCTIONS:
#
# For Turnitin rubrics (.rbc files):
# 1. In Turnitin, create and save a new rubric (or select an existing one)
#    and ensure its Scoring type is set to "Custom".
#
# 2. Download the rubric as a .rbc file from Turnitin. (The Download option is
#    available in the sandwich menu — the three lines menu — for the rubric.)
#
# 3. To convert the .rbc file to Excel for editing:
#      python rubric_converter.py yourrubric.rbc
#    This will produce yourrubric.xlsx.
#
# For IMS specification rubrics (.json files):
# 1. To convert an IMS .json file to Excel for editing:
#      python rubric_converter.py yourrubric.json
#    This will produce yourrubric.xlsx.
#
# Editing the Excel file:
# 1. Edit the Excel file as needed using Microsoft Excel or compatible
#    spreadsheet editor.
#
#    - To specify a criterion title and description: put the criterion title on
#      the first line of the cell in the "Criterion (name and description)"
#      column, and the description (if any) starting on the second line.
#
#    - To specify the description and point value for a cell under a scale,
#      enter the description, then (optionally) the point value in square
#      brackets at the end, e.g.:
#         Adequate analysis [4]
#      or, if only a value is needed:
#         [2]
#      If both are omitted, the cell value is treated as 0.
#
# Converting back from Excel:
# 1. To convert the edited Excel file back to Turnitin format (.rbc):
#      python rubric_converter.py yourrubric.xlsx
#    This will produce yourrubric.rbc (default).
#
# 2. To convert the edited Excel file to IMS format (.json):
#      python rubric_converter.py yourrubric.xlsx -f ims
#    This will produce yourrubric.json.
#
#    You can override the output filename with -o OUTPUT.
#    You can set a new rubric name (when converting Excel to any format) with
#    -r "My Rubric Name".
#
# 3. Upload the .rbc file back into Turnitin (The Upload option is
#    available in the sandwich menu — the three lines menu — for the rubric.)
#    or use the .json file with IMS-compatible systems.
#
# See help with:
#   python rubric_converter.py -h
#
# -----------------------------------------------------------------------------

import pandas as pd
import json
import argparse
import os
import re
import sys
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def generate_id(start):
    current = start
    while True:
        yield current
        current += 1

def parse_desc_value(cell):
    if not isinstance(cell, str) or cell.strip() == "":
        return None, 0
    match = re.match(r"^(.*?)(?:\s*\[(.*?)\])?$", cell.strip())
    if match:
        desc = match.group(1).strip() if match.group(1) else None
        value = match.group(2)
        if value is not None and value != "":
            try:
                value = float(value)
                if value.is_integer():
                    value = int(value)
            except Exception:
                pass
        else:
            value = 0
        if desc == "":
            desc = None
        return desc, value
    return None, 0

def parse_criterion_cell(cell):
    if not isinstance(cell, str) or cell.strip() == "":
        return "", ""
    lines = cell.splitlines()
    name = lines[0].strip() if lines else ""
    desc = "\n".join(lines[1:]).strip() if len(lines) > 1 else ""
    return name, desc

def truncate(val, n):
    val = val or ""
    return val[:n] if len(val) > n else val

def excel_to_rbc(input_excel, output_rbc, rubric_name_override=None):
    base = os.path.basename(input_excel)
    raw_rubric_name = os.path.splitext(base)[0].replace("_", " ")
    rubric_name = rubric_name_override or raw_rubric_name
    rubric_name = truncate(rubric_name, 30)
    truncation_warnings = []

    if rubric_name_override and rubric_name != rubric_name_override:
        truncation_warnings.append(f"Rubric name truncated: '{rubric_name_override}' → '{rubric_name}'")
    elif not rubric_name_override and rubric_name != raw_rubric_name:
        truncation_warnings.append(f"Rubric name truncated: '{raw_rubric_name}' → '{rubric_name}'")

    df = pd.read_excel(input_excel)

    scale_names = [col[:-15] for col in df.columns if col.endswith('(desc [value])')]
    scale_names_unique = []
    for name in scale_names:
        truncated = truncate(name, 25)
        if name != truncated:
            truncation_warnings.append(f"Scale name truncated: '{name}' → '{truncated}'")
        scale_names_unique.append(truncated)
    scale_names = list(dict.fromkeys(scale_names_unique))

    scale_id_gen = generate_id(1_000_000)
    crit_id_gen = generate_id(2_000_000)
    cs_id_gen = generate_id(3_000_000)

    rubric_scales = []
    scale_name_to_id = {}
    for i, scale_name in enumerate(scale_names):
        scale_id = next(scale_id_gen)
        scale_name_to_id[scale_name] = scale_id
        rubric_scales.append({
            "id": scale_id,
            "num": i+1,
            "position": i+1,
            "value": 0,
            "name": scale_name,
            "rubric": 1
        })

    rubric_criteria = []
    rubric_criterion_scales = []
    for idx, row in df.iterrows():
        crit_cell = row['Criterion (name and description)']
        crit_name, crit_desc = parse_criterion_cell(crit_cell)
        raw_crit_name = crit_name
        crit_name = truncate(crit_name, 13)
        if crit_name != raw_crit_name:
            truncation_warnings.append(f"Criterion name truncated: '{raw_crit_name}' → '{crit_name}'")
        crit_id = next(crit_id_gen)

        crit_scales_this = []
        for scale_name in scale_names:
            cs_id = next(cs_id_gen)
            cell = row.get(f"{scale_name} (desc [value])", None)
            desc, value = parse_desc_value(cell)
            crit_scales_this.append(cs_id)
            rubric_criterion_scales.append({
                "criterion": crit_id,
                "scale_value": scale_name_to_id[scale_name],
                "description": desc,
                "value": value,
                "id": cs_id
            })

        rubric_criteria.append({
            "value": 0,
            "id": crit_id,
            "rubric": 1,
            "name": crit_name,
            "description": crit_desc if crit_desc else None,
            "criterion_scales": crit_scales_this,
            "position": idx+1,
            "previous_version": None,
            "num": idx+1
        })

    rubric = [{
        "total_points": None,
        "criterion": [c['id'] for c in rubric_criteria],
        "id": 1,
        "scoring_method": 4,
        "name": rubric_name,
        "distribute_criterion_percentage": 0,
        "rubric_group": None,
        "is_starred": 0,
        "deleted": 0,
        "criterion_scales_all": [cs['id'] for cs in rubric_criterion_scales],
        "scale_values": [s['id'] for s in rubric_scales],
        "papers_scored": 0,
        "owner": 0,
        "cv_loaded": "1",
        "description": None
    }]

    output = {
        "Rubric": rubric,
        "RubricCriterion": rubric_criteria,
        "RubricScale": rubric_scales,
        "RubricCriterionScale": rubric_criterion_scales
    }

    print(f"Converting Excel to RBC format.")
    print(f"Rubric name: {rubric_name}")
    print(f"Number of criteria: {len(rubric_criteria)}")
    print(f"Number of scales: {len(rubric_scales)}")
    print(f"Writing to: {output_rbc}")

    with open(output_rbc, 'w') as f:
        json.dump(output, f, indent=2)

    if truncation_warnings:
        print("WARNING: The following names were truncated to meet length restrictions:", file=sys.stderr)
        for w in truncation_warnings:
            print(w, file=sys.stderr)

def format_desc_value(desc, value):
    if desc and value not in ("", None):
        return f"{desc} [{value}]"
    elif desc:
        return desc
    elif value not in ("", None):
        return f"[{value}]"
    else:
        return ""

def criterion_cell(name, desc):
    if desc and desc.strip():
        return f"{name}\n{desc.strip()}"
    else:
        return name

def is_ims_format(data):
    """Check if the JSON data is in IMS format."""
    if isinstance(data, dict):
        # IMS format has 'CFRubricCriterion' key (IMS Global standard)
        has_cf_rubric_criterion = 'CFRubricCriterion' in data
        # Legacy check for older format: 'criteria' or 'type': 'Rubric'
        has_criteria = 'criteria' in data
        has_type = data.get('type') == 'Rubric' or data.get('@type') == 'Rubric'
        # IMS format doesn't have Turnitin-specific keys
        has_turnitin_keys = 'Rubric' in data or 'RubricCriterion' in data
        return (has_cf_rubric_criterion or has_criteria or has_type) and not has_turnitin_keys
    return False

def ims_to_excel(input_ims, output_excel):
    """Convert IMS format JSON to Excel.
    
    Supports both IMS Global CFRubric format and legacy formats.
    IMS format allows different numbers of levels for different criteria,
    so we need to create a flexible Excel structure.
    """
    from openpyxl.utils import get_column_letter
    
    with open(input_ims, 'r') as f:
        data = json.load(f)
    
    # Extract rubric information - support both CFRubric and legacy formats
    rubric_name = data.get('Title') or data.get('title', 'N/A')
    
    # Check for CFRubricCriterion (IMS Global standard) or legacy 'criteria'
    if 'CFRubricCriterion' in data:
        criteria_data = data['CFRubricCriterion']
    else:
        criteria_data = data.get('criteria', [])
    
    if not criteria_data:
        raise ValueError("No criteria found in IMS format file")
    
    # Find the maximum number of levels across all criteria
    max_levels = 0
    for criterion in criteria_data:
        # Support both CFRubricCriterionLevels and legacy 'levels'
        if 'CFRubricCriterionLevels' in criterion:
            num_levels = len(criterion['CFRubricCriterionLevels'])
        else:
            num_levels = len(criterion.get('levels', []))
        if num_levels > max_levels:
            max_levels = num_levels
    
    # Build the Excel structure with generic level column names
    # Note: Unlike Turnitin format which uses named scales (e.g., "Excellent", "Good"),
    # IMS format supports variable numbers of levels per criterion, so we use
    # generic numbered columns (e.g., "Level 1", "Level 2") to accommodate this flexibility.
    columns = ['Criterion (name and description)']
    for i in range(max_levels):
        columns.append(f"Level {i+1} (desc [value])")
    
    rows = []
    for criterion in criteria_data:
        # Support both CFRubric and legacy formats for criterion description
        crit_name = criterion.get('Description') or criterion.get('title', '')
        crit_desc = criterion.get('description', '')  # CFRubric format typically doesn't have separate description
        row = [criterion_cell(crit_name, crit_desc)]
        
        # Get levels - support both CFRubricCriterionLevels and legacy 'levels'
        if 'CFRubricCriterionLevels' in criterion:
            levels = criterion['CFRubricCriterionLevels']
        else:
            levels = criterion.get('levels', [])
        
        # Add each level for this criterion
        for level in levels:
            # Support both CFRubric format (Description, score) and legacy format (title, description, points)
            desc = level.get('Description') or level.get('description', '')
            points = level.get('score') or level.get('points', 0)
            # Convert score to numeric if it's a string
            if isinstance(points, str):
                try:
                    points = float(points)
                    if points.is_integer():
                        points = int(points)
                except (ValueError, AttributeError):
                    points = 0
            
            # For legacy format, include level title in the description for clarity
            level_title = level.get('title', '')
            if level_title and desc:
                full_desc = f"{level_title}: {desc}"
            elif level_title:
                full_desc = level_title
            else:
                full_desc = desc
            row.append(format_desc_value(full_desc, points))
        
        # Fill remaining columns with empty strings if this criterion has fewer levels
        while len(row) < len(columns):
            row.append('')
        
        rows.append(row)
    
    print(f"Converting IMS format to Excel.")
    print(f"Rubric name: {rubric_name}")
    print(f"Number of criteria: {len(criteria_data)}")
    print(f"Maximum levels per criterion: {max_levels}")
    print(f"Writing to: {output_excel}")
    
    df = pd.DataFrame(rows, columns=columns)
    
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        worksheet = list(writer.sheets.values())[0]
        # Set column widths: first column 5 cm, others 3 cm (1 cm ~ 2.835 units)
        worksheet.column_dimensions[get_column_letter(1)].width = 5 * 2.835  # 5 cm
        for col in range(2, len(columns) + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 3 * 2.835  # 3 cm
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

def excel_to_ims(input_excel, output_ims, rubric_name_override=None, use_cf_format=True):
    """Convert Excel to IMS format JSON.
    
    IMS format allows different numbers of levels for different criteria.
    Only non-empty cells will be converted to levels.
    
    Args:
        input_excel: Path to input Excel file
        output_ims: Path to output JSON file
        rubric_name_override: Optional rubric name override
        use_cf_format: If True, uses CFRubric format (IMS Global standard). 
                       If False, uses legacy format. Default is True.
    """
    import uuid
    from datetime import datetime, timezone
    
    base = os.path.basename(input_excel)
    raw_rubric_name = os.path.splitext(base)[0].replace("_", " ")
    rubric_name = rubric_name_override or raw_rubric_name
    
    df = pd.read_excel(input_excel)
    
    # Extract level column names from column headers
    level_columns = [col for col in df.columns if col.endswith('(desc [value])')]
    
    # Placeholder URI - in production this should be configurable
    base_uri = "https://example.edu/rubric"
    current_time = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S+00:00")
    
    # Build criteria array
    criteria = []
    total_levels = 0
    for idx, row in df.iterrows():
        crit_cell = row['Criterion (name and description)']
        crit_name, crit_desc = parse_criterion_cell(crit_cell)
        
        # Build levels for this criterion - only include non-empty cells
        levels = []
        level_position = 0
        for col_idx, col_name in enumerate(level_columns):
            cell = row.get(col_name, None)
            # Skip empty cells
            if pd.isna(cell) or (isinstance(cell, str) and cell.strip() == ""):
                continue
                
            desc, value = parse_desc_value(cell)
            
            # For CFRubric format, we only use the description (not separate title)
            level_desc = desc if desc else ""
            
            if use_cf_format:
                level = {
                    "score": str(value),
                    "Identifier": str(uuid.uuid4()),
                    "URI": base_uri,
                    "lastChangeDateTime": current_time,
                    "position": level_position,
                    "Description": level_desc
                }
            else:
                # Legacy format with separate title extraction
                level_title = f"Level {level_position + 1}"
                if desc and ":" in desc:
                    parts = desc.split(":", 1)
                    level_title = parts[0].strip()
                    level_desc = parts[1].strip()
                elif desc:
                    desc_words = desc.split()
                    level_title = desc_words[0] if desc_words else f"Level {level_position + 1}"
                
                level = {
                    "id": f"criterion-{idx+1}-level-{level_position+1}",
                    "title": level_title,
                    "description": level_desc,
                    "points": value
                }
            
            levels.append(level)
            level_position += 1
        
        total_levels += len(levels)
        
        if use_cf_format:
            criterion = {
                "Identifier": str(uuid.uuid4()),
                "URI": base_uri,
                "lastChangeDateTime": current_time,
                "position": idx + 1,
                "Description": crit_name,
                "CFRubricCriterionLevels": levels
            }
        else:
            criterion = {
                "id": f"criterion-{idx+1}",
                "title": crit_name,
                "description": crit_desc if crit_desc else "",
                "levels": levels
            }
        
        criteria.append(criterion)
    
    # Build IMS format output
    if use_cf_format:
        # IMS Global CFRubric format
        output = {
            "description": "",
            "Identifier": str(uuid.uuid4()),
            "URI": base_uri,
            "Title": rubric_name,
            "lastChangeDateTime": current_time,
            "CFRubricCriterion": criteria
        }
    else:
        # Legacy format
        # Note: Using example.edu as a placeholder domain for rubric IDs.
        # In production use, this should be replaced with the actual institutional domain.
        output = {
            "@context": "http://purl.imsglobal.org/ctx/caliper/v1p2",
            "type": "Rubric",
            "id": f"https://example.edu/rubrics/{rubric_name.lower().replace(' ', '-')}",
            "title": rubric_name,
            "description": "",
            "criteria": criteria
        }
    
    print(f"Converting Excel to IMS format ({'CFRubric' if use_cf_format else 'legacy'}).")
    print(f"Rubric name: {rubric_name}")
    print(f"Number of criteria: {len(criteria)}")
    print(f"Total levels across all criteria: {total_levels}")
    print(f"Writing to: {output_ims}")
    
    with open(output_ims, 'w') as f:
        json.dump(output, f, indent=2)

def rbc_to_excel(input_rbc, output_excel):
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook

    with open(input_rbc, 'r') as f:
        data = json.load(f)
    
    # Check if this is IMS format
    if is_ims_format(data):
        # Close the file and call ims_to_excel
        return ims_to_excel(input_rbc, output_excel)

    criteria = {c['id']: c for c in data['RubricCriterion']}
    scales = {s['id']: s for s in data['RubricScale']}
    criterion_scale_map = {}

    for cs in data['RubricCriterionScale']:
        crit_id = cs['criterion']
        scale_id = cs['scale_value']
        if crit_id not in criterion_scale_map:
            criterion_scale_map[crit_id] = {}
        criterion_scale_map[crit_id][scale_id] = cs

    columns = ['Criterion (name and description)']
    for scale in sorted(scales.values(), key=lambda x: x['position']):
        columns.append(f"{scale['name']} (desc [value])")

    rows = []
    for crit_id, crit in criteria.items():
        row = [criterion_cell(crit['name'], crit.get('description', ''))]
        for scale in sorted(scales.values(), key=lambda x: x['position']):
            cs = criterion_scale_map.get(crit_id, {}).get(scale['id'], {})
            desc = cs.get('description', '')
            value = cs.get('value', '')
            row.append(format_desc_value(desc, value))
        rows.append(row)

    rubric_name = data['Rubric'][0].get('name', 'N/A') if data.get('Rubric') and isinstance(data['Rubric'], list) else 'N/A'
    print(f"Converting RBC/JSON to Excel format.")
    print(f"Rubric name: {rubric_name}")
    print(f"Number of criteria: {len(criteria)}")
    print(f"Number of scales: {len(scales)}")
    print(f"Writing to: {output_excel}")

    df = pd.DataFrame(rows, columns=columns)

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        worksheet = list(writer.sheets.values())[0]
        # Set column widths: first column 5 cm, others 3 cm (1 cm ~ 2.835 units)
        worksheet.column_dimensions[get_column_letter(1)].width = 5 * 2.835  # 5 cm
        for col in range(2, len(columns) + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 3 * 2.835  # 3 cm
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Convert between Turnitin rubric Excel (.xlsx) and RBC/JSON (.rbc/.json) "
            "formats.\n\n"
            "USAGE INSTRUCTIONS:\n"
            "\n"
            "1. In Turnitin, create and save a new rubric (or select an existing one)\n"
            "   and ensure its Scoring type is set to \"Custom\".\n"
            "\n"
            "2. Download the rubric as a .rbc file from Turnitin. (The Download option is\n"
            "   available in the sandwich menu — the three lines menu — for the rubric.)\n"
            "\n"
            "3. To convert the .rbc file to Excel for editing:\n"
            "     python rubric_converter.py yourrubric.rbc\n"
            "   This will produce yourrubric.xlsx.\n"
            "\n"
            "4. Edit the Excel file as needed using Microsoft Excel or compatible\n"
            "   spreadsheet editor.\n"
            "\n"
            "   - To specify a criterion title and description: put the criterion title on\n"
            "     the first line of the cell in the \"Criterion (name and description)\"\n"
            "     column, and the description (if any) starting on the second line.\n"
            "\n"
            "   - To specify the description and point value for a cell under a scale,\n"
            "     enter the description, then (optionally) the point value in square\n"
            "     brackets at the end, e.g.:\n"
            "        Adequate analysis [4]\n"
            "     or, if only a value is needed:\n"
            "        [2]\n"
            "     If both are omitted, the cell value is treated as 0.\n"
            "\n"
            "5. To convert the edited Excel file back to .rbc:\n"
            "     python rubric_converter.py yourrubric.xlsx\n"
            "   This will produce yourrubric.rbc.\n"
            "   You can override the output filename with -o OUTPUT.\n"
            "   You can set a new rubric name (when converting Excel to RBC) with\n"
            "   -r \"My Rubric Name\".\n"
            "\n"
            "6. Upload the .rbc file back into Turnitin. (The Upload option is\n"
            "   available in the sandwich menu — the three lines menu — for the rubric.)\n"
            "\n"
            "See help with:\n"
            "  python rubric_converter.py -h"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("input_file", help="Path to the input file (.xlsx, .rbc, or .json)")
    parser.add_argument("-o", "--output_file", help="Optional output file name")
    parser.add_argument("-r", "--rubric-name", help="Rubric name (overrides name from file name, Excel→JSON only)")
    parser.add_argument("-f", "--format", choices=["turnitin", "ims"], 
                       help="Output format when converting from Excel (default: turnitin)")
    args = parser.parse_args()

    input_ext = os.path.splitext(args.input_file)[1].lower()
    if input_ext == ".xlsx":
        # Determine output format
        output_format = args.format or "turnitin"
        
        if args.output_file:
            output_file = args.output_file
        else:
            base, _ = os.path.splitext(args.input_file)
            if output_format == "ims":
                output_file = base + ".json"
            else:
                output_file = base + ".rbc"
        
        # Convert based on format
        if output_format == "ims":
            excel_to_ims(args.input_file, output_file, rubric_name_override=args.rubric_name)
        else:
            excel_to_rbc(args.input_file, output_file, rubric_name_override=args.rubric_name)
    elif input_ext in (".rbc", ".json"):
        if args.output_file:
            output_excel = args.output_file
        else:
            base, _ = os.path.splitext(args.input_file)
            output_excel = base + ".xlsx"
        rbc_to_excel(args.input_file, output_excel)
    else:
        print("Unrecognized input file extension. Please provide a .xlsx, .rbc, or .json file.", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()